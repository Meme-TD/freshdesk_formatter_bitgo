import os
import re
import sys
import csv
import time
import json
import pathlib
import requests
from collections import Counter
from bs4 import BeautifulSoup
from openpyxl import load_workbook


FRESHDESK_DOMAIN = "bitgo"
API_KEY = os.environ.get("FRESHDESK_API_KEY", "")
XLSX_PATH = "FD_canned_response_and_KB_articles_clean-up_-_March_2026__2_.xlsx"
SHEET_NAME = "All KB Articles"

MODE = "dry_run"
INSPECT_COUNT = 3
LIMIT = None
ONLY_ARTICLE_IDS = []       
REQUEST_PAUSE = 0.6

REQUIRE_PUBLISHED = True
EXCLUDE_FOLDER_SUBSTR = "deprecat"

# Colors / sizes
TITLE_BLUE_OTHER = "#1647DB" #rgb(22,71,219)
LINK_COLOR_UG    = "#173ECA" #rgb(23,62,202)
BLACK            = "#000000"
FONT_FAMILY      = "Arial, sans-serif"
TITLE_SIZE       = "30px"
SUBTITLE_SIZE    = "24px"
BODY_SIZE        = "16px"

REMOVE_EMPTY_SPACERS  = True
SUBTITLE_MAX_CHARS    = 120    
SUBTITLE_SIZE_MARGIN_PX = 0.5  

#FRESHDESK APIIIIIIII
BASE = f"https://{FRESHDESK_DOMAIN}.freshdesk.com/api/v2"
AUTH = (API_KEY, "X")


def fd_get_article(aid):
    return _request("GET", f"{BASE}/solutions/articles/{aid}")


def fd_update_article(aid, html):
    return _request("PUT", f"{BASE}/solutions/articles/{aid}", json={"description": html})


def _request(method, url, **kw):
    for _ in range(6):
        r = requests.request(method, url, auth=AUTH, timeout=30, **kw)
        if r.status_code == 429:
            wait = int(r.headers.get("Retry-After", "10"))
            print(f"  rate limited, sleeping {wait}s"); time.sleep(wait); continue
        r.raise_for_status(); return r.json()
    raise RuntimeError(f"Too many retries for {url}")

#spreadsheet garbage
def read_articles():
    ws = load_workbook(XLSX_PATH)[SHEET_NAME]
    out = []
    for r in range(2, ws.max_row + 1):
        category, folder, title = ws.cell(r, 2).value, ws.cell(r, 3).value, ws.cell(r, 4).value
        cell = ws.cell(r, 5)
        link = cell.value or (cell.hyperlink.target if cell.hyperlink else None)
        published = ws.cell(r, 6).value
        if category is None and title is None and link is None:
            continue
        if REQUIRE_PUBLISHED and str(published).strip().lower() != "yes":
            continue
        if folder and EXCLUDE_FOLDER_SUBSTR in str(folder).lower():
            continue
        m = re.search(r"/articles/(\d+)", str(link)) if link else None
        if not m:
            continue
        out.append({"id": m.group(1), "title": (title or "").strip(),
                    "category": str(category).strip(),
                    "is_user_guide": str(category).strip().lower() == "bitgo user guide",
                    "row": r})
    return out


#style
def parse_style(s):
    d = {}
    for part in (s or "").split(";"):
        if ":" in part:
            k, v = part.split(":", 1)
            d[k.strip().lower()] = v.strip()
    return d


def dump_style(d):
    return "; ".join(f"{k}: {v}" for k, v in d.items())


def set_styles(el, **props):
    d = parse_style(el.get("style", ""))
    for k, v in props.items():
        d[k.replace("_", "-")] = v
    el["style"] = dump_style(d)


def norm(t):
    return re.sub(r"\s+", " ", (t or "")).strip().lower()


def size_to_px(v):
    m = re.match(r"([\d.]+)\s*(px|pt)?", (v or "").strip())
    if not m:
        return None
    num = float(m.group(1)); unit = m.group(2) or "px"
    return num * (96.0 / 72.0) if unit == "pt" else num


def find_title_element(soup, article_title):
    want = norm(article_title)
    if want:
        for tags in (["h1", "h2", "h3", "h4", "h5", "h6"],
                     ["p", "strong", "b", "span", "div"]):
            for el in soup.find_all(tags):
                if norm(el.get_text()) == want:
                    return el
    return soup.find(["h1", "h2", "h3"])


def center_image(img):
    # clear Froala float-left and center as a block; harmless inside flex wrappers
    set_styles(img, float="none", display="block",
               margin_left="auto", margin_right="auto")


def override_descendant_color(el, color):
    for d in el.find_all(True):
        st = parse_style(d.get("style", ""))
        if "color" in st:
            st["color"] = color
            d["style"] = dump_style(st)


def restyle_ug_block(block, size, color, bold):
    """Arial + size (+ optional color/bold), normalizing nested spans.
       color=None -> leave colors (title keeps its color)
       bold=True  -> force bold (strip descendant weights so block wins)
       bold=False -> keep descendant weights (preserve inline bold body text)"""
    set_styles(block, font_family=FONT_FAMILY, font_size=size)
    if bold:
        set_styles(block, font_weight="bold")
    if color is not None:
        set_styles(block, color=color)
    for d in block.find_all(True):
        st = parse_style(d.get("style", ""))
        st.pop("font-family", None)
        st.pop("font-size", None)
        if color is not None:
            st.pop("color", None)
        if bold:
            st.pop("font-weight", None)
        if st:
            d["style"] = dump_style(st)
        elif "style" in d.attrs:
            del d["style"]

HEADINGS = ["h1", "h2", "h3", "h4", "h5", "h6"]


def is_spacer(el):
    return (el.name in ["p"] + HEADINGS
            and not el.get_text(strip=True)
            and el.find("img") is None)


def block_min_text_size_px(el):
    sizes = []
    for node in [el] + el.find_all(True):
        st = parse_style(node.get("style", ""))
        if "font-size" in st and node.get_text(strip=True):
            px = size_to_px(st["font-size"])
            if px:
                sizes.append(px)
    return min(sizes) if sizes else None


#non user guide
def transform_other(html, article_title):
    soup = BeautifulSoup(html or "", "html.parser")
    title_el = find_title_element(soup, article_title)
    if title_el is not None:
        set_styles(title_el, color=TITLE_BLUE_OTHER)
        override_descendant_color(title_el, TITLE_BLUE_OTHER)
    for img in soup.find_all("img"):
        center_image(img)
    report = {"title": title_el.get_text(strip=True) if title_el else None, "subtitles": []}
    return str(soup), (title_el is not None), report


#user guide
def transform_user_guide(html, article_title):
    soup = BeautifulSoup(html or "", "html.parser")

    if REMOVE_EMPTY_SPACERS:
        for el in soup.find_all(["p"] + HEADINGS):
            if is_spacer(el):
                el.decompose()

    title_el = find_title_element(soup, article_title)
    title_id = id(title_el) if title_el is not None else None

    # candidate content blocks
    blocks = [el for el in soup.find_all(["p", "li"] + HEADINGS)
              if el.get_text(strip=True) or el.find("img")]

    # body size = most common representative size among non-title blocks
    sizes = []
    for el in blocks:
        if id(el) == title_id:
            continue
        s = block_min_text_size_px(el)
        if s:
            sizes.append(round(s, 1))
    body_px = Counter(sizes).most_common(1)[0][0] if sizes else None

    # classify subtitles
    subtitle_ids = set()
    subtitle_texts = []
    for el in blocks:
        if id(el) == title_id:
            continue
        is_heading = el.name in HEADINGS
        looks_like_subtitle = False
        if not is_heading and el.name != "li":
            s = block_min_text_size_px(el)
            txt = el.get_text(strip=True)
            if (s is not None and body_px is not None
                    and s > body_px + SUBTITLE_SIZE_MARGIN_PX
                    and 0 < len(txt) <= SUBTITLE_MAX_CHARS
                    and el.find("br") is None):
                looks_like_subtitle = True
        if is_heading or looks_like_subtitle:
            subtitle_ids.add(id(el))
            subtitle_texts.append(el.get_text(strip=True))

    # restyle: title, subtitles, body
    if title_el is not None:
        restyle_ug_block(title_el, TITLE_SIZE, color=None, bold=True)
    for el in blocks:
        if id(el) in subtitle_ids:
            restyle_ug_block(el, SUBTITLE_SIZE, color=BLACK, bold=True)
    for el in blocks:
        if id(el) == title_id or id(el) in subtitle_ids:
            continue
        restyle_ug_block(el, BODY_SIZE, color=BLACK, bold=False)

    for h in soup.find_all(HEADINGS):
        h.name = "p"

    for a in soup.find_all("a"):
        set_styles(a, color=LINK_COLOR_UG)
        override_descendant_color(a, LINK_COLOR_UG)

    for img in soup.find_all("img"):                       # 1 after each image
        img.insert_after(soup.new_tag("br"))

    for el in soup.find_all("p"):                          # 2 before each subheader
        if id(el) in subtitle_ids and el.find_previous_sibling() is not None:
            el.insert_before(soup.new_tag("br"))
            el.insert_before(soup.new_tag("br"))

    for p in [e for e in soup.find_all("p")                # 1 between body paragraphs
              if id(e) != title_id and id(e) not in subtitle_ids]:
        nxt = p.find_next_sibling()
        if getattr(nxt, "name", None) == "p" and id(nxt) not in subtitle_ids \
                and id(nxt) != title_id:
            p.insert_after(soup.new_tag("br"))

    for _ in range(3):                                     # 3 at end
        soup.append(soup.new_tag("br"))

    report = {"title": title_el.get_text(strip=True) if title_el else None,
              "subtitles": subtitle_texts, "body_px": body_px}
    return str(soup), (title_el is not None), report



def main():
    global MODE
    if len(sys.argv) > 1:
        MODE = sys.argv[1]
    if MODE != "inspect" and not API_KEY:
        sys.exit("ERROR: set the FRESHDESK_API_KEY environment variable first.")

    articles = read_articles()
    if ONLY_ARTICLE_IDS:
        keep = set(ONLY_ARTICLE_IDS)
        articles = [a for a in articles if a["id"] in keep]
    if LIMIT:
        articles = articles[:LIMIT]

    ug = [a for a in articles if a["is_user_guide"]]
    other = [a for a in articles if not a["is_user_guide"]]
    print(f"MODE={MODE}  user-guide={len(ug)}  other={len(other)}  total={len(articles)}")

    if MODE == "inspect":
        outdir = pathlib.Path("fd_inspect"); outdir.mkdir(exist_ok=True)
        for a in ug[:INSPECT_COUNT] + other[:INSPECT_COUNT]:
            data = fd_get_article(a["id"]); tag = "UG" if a["is_user_guide"] else "OTHER"
            (outdir / f"{tag}_{a['id']}.html").write_text(data.get("description") or "", encoding="utf-8")
            print(f"  dumped {tag} {a['id']}  title={data.get('title')!r}")
            time.sleep(REQUEST_PAUSE)
        print(f"Raw HTML in {outdir}/")
        return

    dryroot = pathlib.Path("fd_dryrun")
    report_rows = []
    if MODE == "dry_run":
        dryroot.mkdir(exist_ok=True)

    missing = []
    for i, a in enumerate(articles, 1):
        try:
            data = fd_get_article(a["id"])
        except Exception as e:
            print(f"  [{i}/{len(articles)}] GET {a['id']} FAILED: {e}"); continue
        html = data.get("description") or ""
        fn = transform_user_guide if a["is_user_guide"] else transform_other
        new_html, found, rep = fn(html, a["title"])
        if not found:
            missing.append(a["id"])
        tag = "UG" if a["is_user_guide"] else "OTHER"

        if MODE == "dry_run":
            (dryroot / f"{tag}_{a['id']}.before.html").write_text(html, encoding="utf-8")
            (dryroot / f"{tag}_{a['id']}.after.html").write_text(new_html, encoding="utf-8")
            report_rows.append({"id": a["id"], "group": tag, "sheet_title": a["title"],
                                "detected_title": rep["title"],
                                "n_subtitles": len(rep["subtitles"]),
                                "subtitles": " | ".join(rep["subtitles"])})
            print(f"  [{i}/{len(articles)}] {tag} {a['id']} title-found={found} "
                  f"subtitles={len(rep['subtitles'])}")
        elif MODE == "apply":
            try:
                fd_update_article(a["id"], new_html)
                print(f"  [{i}/{len(articles)}] updated {a['id']}")
            except Exception as e:
                print(f"  [{i}/{len(articles)}] PUT {a['id']} FAILED: {e}")
        time.sleep(REQUEST_PAUSE)

    if MODE == "dry_run" and report_rows:
        with open(dryroot / "_classification.csv", "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=list(report_rows[0].keys()))
            w.writeheader(); w.writerows(report_rows)
        print(f"\nBefore/after HTML + _classification.csv in {dryroot}/ — review before apply.")
    if missing:
        print(f"\nWARNING: in-body title not found for {len(missing)} articles: {missing}")


if __name__ == "__main__":
    main()
